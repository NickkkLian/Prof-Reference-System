import os, re
from openpyxl import load_workbook
import config, database as db

REQUIRED_COLS_WITH_META    = {"student #", "first name", "last name", "%abs", "grade"}
REQUIRED_COLS_WITHOUT_META = {"student #", "first name", "last name", "%abs", "grade"}

# Accepted column names for absence rate
ABS_COL_NAMES = {"%abs", "abs", "absence", "absence %", "absent %",
                 "% abs", "abs%", "absences", "abs rate", "absence rate"}

def _build_id(course, section, year, term, sno):
    def c(v): return re.sub(r"[^A-Z0-9]", "", str(v).strip().upper())
    return "_".join(c(x) for x in [course, section, year, term, sno])

def parse_roster_file(filepath, course=None, section=None, year=None, term=None):
    """
    Parse a roster Excel file.
    If course/section/year/term are provided, they override any columns in the file.
    If not provided, the file must contain those columns.
    """
    wb = load_workbook(filepath, data_only=True)
    ws = wb["Roster"] if "Roster" in wb.sheetnames else wb.worksheets[0]

    raw = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
    idx = {h: i for i, h in enumerate(raw) if h}

    # Find the absence column using any accepted name
    abs_col = next((h for h in idx if h in ABS_COL_NAMES), None)
    if abs_col:
        idx["%abs"] = idx[abs_col]  # normalise to %abs key

    # Determine mode: metadata from form or from file
    meta_from_form = all([course, section, year, term])

    if meta_from_form:
        required = {"student #", "first name", "last name", "%abs"}
        missing  = required - set(idx.keys())
        if missing:
            raise ValueError(
                f"Missing columns: {missing}. "
                f"File must contain: Student #, First Name, Last Name, "
                f"and an absence column (e.g. Abs, %Abs, Absence). "
                f"Grade column is optional."
            )
    else:
        required = {"student #", "first name", "last name", "year",
                    "term", "course", "section", "%abs"}
        missing  = required - set(idx.keys())
        if missing:
            raise ValueError(
                f"Missing columns: {missing}. "
                f"Either enter Course/Section/Year/Term in the form, "
                f"or include those columns in the file."
            )

    records = []
    for r in range(2, ws.max_row + 1):
        vals = [ws.cell(row=r, column=c + 1).value for c in range(ws.max_column)]

        def g(col):
            return vals[idx[col]] if idx.get(col) is not None and idx[col] < len(vals) else None

        sno = g("student #")
        if sno is None or str(sno).strip() == "":
            continue

        fn  = str(g("first name") or "").strip()
        ln  = str(g("last name")  or "").strip()
        sno = str(sno).strip()

        # Use form metadata if provided, otherwise read from file
        row_course  = course  if meta_from_form else str(g("course")  or "").strip()
        row_section = section if meta_from_form else str(g("section") or "").strip()
        row_year    = year    if meta_from_form else str(g("year")    or "").strip()
        row_term    = term    if meta_from_form else str(g("term")    or "").strip()

        abs_raw = g("%abs")
        try:
            av = float(abs_raw) if abs_raw not in (None, "") else 0.0
            if av > 1.0: av /= 100.0
        except (ValueError, TypeError):
            av = 0.0

        grade_raw = g("grade")
        try:
            grade = float(grade_raw) if grade_raw not in (None, "") else None
        except (ValueError, TypeError):
            grade = None

        records.append({
            "identifier":          _build_id(row_course, row_section, row_year, row_term, sno),
            "course":              row_course,
            "section":             row_section,
            "year":                row_year,
            "term":                row_term,
            "student_number":      sno,
            "student_name":        f"{fn} {ln}".strip(),
            "first_name":          fn,
            "last_name":           ln,
            "absence_rate_pct":    round(av * 100, 2),
            "attendance_rate_pct": round((1.0 - av) * 100, 2),
            "grade":               grade,
            "source_file":         os.path.basename(filepath),
        })
    return records


def import_file(filepath, verbose=True,
                course=None, section=None, year=None, term=None):
    records = parse_roster_file(filepath, course=course, section=section,
                                year=year, term=term)
    added   = db.upsert_many(records)
    if verbose:
        print(f"  ✔ '{os.path.basename(filepath)}' — {len(records)} records, {added} new")
    return {"total": len(records), "new": added}


def import_all(verbose=True):
    files = [f for f in os.listdir(config.ATTENDANCE_INPUT_DIR)
             if f.lower().endswith(".xlsx")]
    results = []
    for fname in files:
        try:
            r = import_file(os.path.join(config.ATTENDANCE_INPUT_DIR, fname), verbose)
            r["file"]   = fname
            r["status"] = "ok"
        except Exception as e:
            r = {"file": fname, "total": 0, "new": 0, "status": f"error: {e}"}
            if verbose:
                print(f"  ✘ '{fname}': {e}")
        results.append(r)
    return results
