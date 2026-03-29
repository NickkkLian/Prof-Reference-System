"""
database.py
-----------
All SQLite database operations for the Professor Reference System.
The database is a single .db file stored in the Drive folder.
"""

import sqlite3
import os
from contextlib import contextmanager
from datetime import datetime, timezone, timedelta
import config

def _vancouver_now_str() -> str:
    """Return current Vancouver time as string for DB storage (PDT = UTC-7)."""
    van_time = datetime.now(timezone.utc) + timedelta(hours=-7)
    return van_time.strftime("%Y-%m-%d %H:%M")

# ---------------------------------------------------------------------------
# Connection
# ---------------------------------------------------------------------------

@contextmanager
def _conn():
    """Context manager — always commits on success, rolls back on error."""
    con = sqlite3.connect(config.DB_PATH, timeout=30)
    con.row_factory = sqlite3.Row
    con.execute("PRAGMA foreign_keys=ON")
    try:
        yield con
        con.commit()
    except Exception:
        con.rollback()
        raise
    finally:
        con.close()


# ---------------------------------------------------------------------------
# Schema creation
# ---------------------------------------------------------------------------

def init_db():
    """Create tables if they don't exist. Safe to call every session."""
    with _conn() as con:
        con.executescript("""
            CREATE TABLE IF NOT EXISTS enrolments (
                identifier        TEXT PRIMARY KEY,
                course            TEXT NOT NULL,
                section           TEXT NOT NULL,
                year              TEXT NOT NULL,
                term              TEXT NOT NULL,
                student_number    TEXT NOT NULL,
                student_name      TEXT,
                first_name        TEXT,
                last_name         TEXT,
                absence_rate_pct  REAL,
                attendance_rate_pct REAL,
                grade             REAL,
                source_file       TEXT,
                updated_at        TEXT DEFAULT (datetime('now'))
            );

            CREATE INDEX IF NOT EXISTS idx_student_number
                ON enrolments(student_number);

            CREATE INDEX IF NOT EXISTS idx_course
                ON enrolments(course, section, year, term);

            CREATE TABLE IF NOT EXISTS eligible_students (
                student_number      TEXT PRIMARY KEY,
                student_name        TEXT,
                overall_grade_pct   REAL,
                grade_source        TEXT,
                eligible_identifiers TEXT,
                attendance_rates    TEXT,
                transcript_file     TEXT,
                letter_file         TEXT,
                checked_at          TEXT DEFAULT (datetime('now')),
                reference_done      INTEGER DEFAULT 0,
                reference_done_at   TEXT
            );
        """)
    # Migration: add columns if they don't exist yet (for existing databases)
    try:
        con = sqlite3.connect(config.DB_PATH)
        con.execute("ALTER TABLE eligible_students ADD COLUMN reference_done INTEGER DEFAULT 0")
        con.execute("ALTER TABLE eligible_students ADD COLUMN reference_done_at TEXT")
        con.commit()
        con.close()
    except Exception:
        pass  # Columns already exist — safe to ignore
    print(f"✅ Database ready: {config.DB_PATH}")


# ---------------------------------------------------------------------------
# Enrolments — write
# ---------------------------------------------------------------------------

def upsert_enrolment(record: dict):
    """
    Insert or update one enrolment record.
    If the identifier already exists, only blank/null fields are overwritten.
    """
    with _conn() as con:
        existing = con.execute(
            "SELECT * FROM enrolments WHERE identifier = ?",
            (record["identifier"],)
        ).fetchone()

        if existing is None:
            con.execute("""
                INSERT INTO enrolments (
                    identifier, course, section, year, term,
                    student_number, student_name, first_name, last_name,
                    absence_rate_pct, attendance_rate_pct, grade, source_file
                ) VALUES (
                    :identifier, :course, :section, :year, :term,
                    :student_number, :student_name, :first_name, :last_name,
                    :absence_rate_pct, :attendance_rate_pct, :grade, :source_file
                )
            """, record)
        else:
            # Only fill in fields that are currently NULL
            updates = []
            params  = []
            for col in ["student_name", "first_name", "last_name",
                        "absence_rate_pct", "attendance_rate_pct",
                        "grade", "source_file"]:
                if existing[col] is None and record.get(col) is not None:
                    updates.append(f"{col} = ?")
                    params.append(record[col])
            if updates:
                params.append(record["identifier"])
                con.execute(
                    f"UPDATE enrolments SET {', '.join(updates)}, "
                    f"updated_at = datetime('now') WHERE identifier = ?",
                    params
                )


def upsert_many(records: list[dict]) -> int:
    """Upsert a list of records. Returns number of new rows inserted."""
    before = count_enrolments()
    for r in records:
        upsert_enrolment(r)
    after = count_enrolments()
    return after - before


# ---------------------------------------------------------------------------
# Enrolments — read
# ---------------------------------------------------------------------------

def count_enrolments() -> int:
    with _conn() as con:
        return con.execute("SELECT COUNT(*) FROM enrolments").fetchone()[0]


def get_all_enrolments() -> list[dict]:
    with _conn() as con:
        rows = con.execute(
            "SELECT * FROM enrolments ORDER BY course, section, year, term, student_number"
        ).fetchall()
        return [dict(r) for r in rows]


def get_enrolments_by_student(student_number: str) -> list[dict]:
    with _conn() as con:
        rows = con.execute(
            "SELECT * FROM enrolments WHERE UPPER(student_number) = UPPER(?)",
            (student_number,)
        ).fetchall()
        return [dict(r) for r in rows]


def get_identifiers_by_student(student_number: str) -> list[str]:
    rows = get_enrolments_by_student(student_number)
    return [r["identifier"] for r in rows]


def get_attendance_rate(identifier: str) -> float | None:
    with _conn() as con:
        row = con.execute(
            "SELECT attendance_rate_pct FROM enrolments WHERE identifier = ?",
            (identifier,)
        ).fetchone()
        return float(row[0]) if row and row[0] is not None else None


def get_grades_by_student(student_number: str) -> dict[str, float | None]:
    rows = get_enrolments_by_student(student_number)
    return {r["identifier"]: r["grade"] for r in rows}


def get_student_name(student_number: str) -> str | None:
    with _conn() as con:
        row = con.execute(
            "SELECT student_name FROM enrolments "
            "WHERE UPPER(student_number) = UPPER(?) AND student_name IS NOT NULL "
            "LIMIT 1",
            (student_number,)
        ).fetchone()
        return row[0] if row else None


def student_exists(student_number: str) -> bool:
    with _conn() as con:
        row = con.execute(
            "SELECT 1 FROM enrolments WHERE UPPER(student_number) = UPPER(?)",
            (student_number,)
        ).fetchone()
        return row is not None


def delete_by_source_file(filename: str) -> dict:
    """
    Delete all enrolments from a specific source file.
    Also removes eligible students who have no remaining enrolments.
    Returns counts of what was deleted.
    """
    with _conn() as con:
        # Get student numbers affected by this file
        affected = con.execute(
            "SELECT DISTINCT student_number FROM enrolments WHERE source_file = ?",
            (filename,)
        ).fetchall()
        affected_students = [r[0] for r in affected]

        # Delete enrolments from this file
        cur = con.execute(
            "DELETE FROM enrolments WHERE source_file = ?", (filename,)
        )
        deleted_enrolments = cur.rowcount

        # For each affected student, check if they still have enrolments
        removed_eligible = 0
        for sno in affected_students:
            remaining = con.execute(
                "SELECT COUNT(*) FROM enrolments WHERE student_number = ?", (sno,)
            ).fetchone()[0]
            if remaining == 0:
                # No enrolments left — remove from eligible too
                cur2 = con.execute(
                    "DELETE FROM eligible_students WHERE student_number = ?", (sno,)
                )
                removed_eligible += cur2.rowcount

    return {
        "deleted_enrolments": deleted_enrolments,
        "removed_eligible":   removed_eligible,
        "affected_students":  len(affected_students),
    }


def get_source_files() -> list[dict]:
    """Return list of source files with their enrolment counts."""
    with _conn() as con:
        rows = con.execute("""
            SELECT source_file,
                   COUNT(*) as enrolment_count,
                   COUNT(DISTINCT student_number) as student_count,
                   MAX(updated_at) as last_updated
            FROM enrolments
            WHERE source_file IS NOT NULL AND source_file != ''
            GROUP BY source_file
            ORDER BY source_file
        """).fetchall()
        return [dict(r) for r in rows]


def get_summary_stats() -> dict:
    with _conn() as con:
        total = con.execute("SELECT COUNT(*) FROM enrolments").fetchone()[0]
        students = con.execute(
            "SELECT COUNT(DISTINCT student_number) FROM enrolments"
        ).fetchone()[0]
        courses = con.execute(
            "SELECT DISTINCT course FROM enrolments ORDER BY course"
        ).fetchall()
        return {
            "total_enrolments": total,
            "unique_students":  students,
            "courses":          [r[0] for r in courses],
        }


# ---------------------------------------------------------------------------
# Eligible students — write
# ---------------------------------------------------------------------------

def upsert_eligible(student_number: str, student_name: str,
                    grade: float | None, grade_source: str,
                    eligible_identifiers: list[str],
                    attendance_rates: dict[str, float | None],
                    transcript_file: str = "",
                    letter_file: str = ""):
    att_str  = "; ".join(
        f"{iid}:{rate:.1f}%" for iid, rate in attendance_rates.items()
        if rate is not None
    )
    iids_str = "; ".join(eligible_identifiers)
    now_van  = _vancouver_now_str()

    with _conn() as con:
        con.execute("""
            INSERT INTO eligible_students (
                student_number, student_name, overall_grade_pct, grade_source,
                eligible_identifiers, attendance_rates,
                transcript_file, letter_file, checked_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            ON CONFLICT(student_number) DO UPDATE SET
                student_name        = excluded.student_name,
                overall_grade_pct   = excluded.overall_grade_pct,
                grade_source        = excluded.grade_source,
                eligible_identifiers = excluded.eligible_identifiers,
                attendance_rates    = excluded.attendance_rates,
                transcript_file     = excluded.transcript_file,
                letter_file         = excluded.letter_file,
                checked_at          = excluded.checked_at
        """, (student_number, student_name, grade, grade_source,
              iids_str, att_str, transcript_file, letter_file, now_van))


# ---------------------------------------------------------------------------
# Eligible students — read
# ---------------------------------------------------------------------------

def toggle_reference_done(student_number: str) -> bool:
    """Toggle reference_done for a student. Returns new state (True=done)."""
    with _conn() as con:
        row = con.execute(
            "SELECT reference_done FROM eligible_students WHERE student_number = ?",
            (student_number,)
        ).fetchone()
        if row is None:
            return False
        new_val = 0 if row[0] else 1
        now     = _vancouver_now_str() if new_val else None
        con.execute(
            "UPDATE eligible_students SET reference_done=?, reference_done_at=? "
            "WHERE student_number=?",
            (new_val, now, student_number)
        )
        return bool(new_val)


def delete_eligible(student_number: str):
    """Remove a student from the eligible list."""
    with _conn() as con:
        con.execute(
            "DELETE FROM eligible_students WHERE student_number = ?",
            (student_number,)
        )


def get_all_eligible() -> list[dict]:
    with _conn() as con:
        rows = con.execute(
            "SELECT * FROM eligible_students ORDER BY checked_at DESC"
        ).fetchall()
        return [dict(r) for r in rows]


def count_eligible() -> int:
    with _conn() as con:
        return con.execute("SELECT COUNT(*) FROM eligible_students").fetchone()[0]


# ---------------------------------------------------------------------------
# Export to Excel (for download / professor view)
# ---------------------------------------------------------------------------

def export_to_excel():
    """Write current DB contents to combined_attendance.xlsx and eligible_students.xlsx."""
    import pandas as pd
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    def _style(ws, n_cols):
        fill_h = PatternFill("solid", start_color="1F4E79")
        font_h = Font(bold=True, color="FFFFFF", name="Arial", size=10)
        for c in range(1, n_cols + 1):
            cell = ws.cell(row=1, column=c)
            cell.fill = fill_h
            cell.font = font_h
            cell.alignment = Alignment(horizontal="center", vertical="center")
        thin = Side(style="thin", color="D9D9D9")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.border = border
                cell.font = Font(name="Arial", size=10)
                fill_color = "DCE6F1" if cell.row % 2 == 0 else "FFFFFF"
                cell.fill = PatternFill("solid", start_color=fill_color)
        ws.freeze_panes = "A2"

    # Enrolments
    rows = get_all_enrolments()
    if rows:
        df = pd.DataFrame(rows)
        df.columns = [c.replace("_", " ").title() for c in df.columns]
        with pd.ExcelWriter(config.COMBINED_EXCEL, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name="Roster", index=False)
            summary_cols = ["Identifier","Course","Section","Year","Term",
                           "Student Number","Student Name",
                           "Attendance Rate Pct","Grade"]
            avail = [c for c in summary_cols if c.replace(" ","_").lower()
                     in [col.replace(" ","_").lower() for col in df.columns]]
            df2 = df[[c for c in df.columns
                      if any(c.lower().replace(" ","_") ==
                             s.lower().replace(" ","_") for s in summary_cols)]]
            df2.to_excel(writer, sheet_name="Summary", index=False)
        # Apply styling
        wb = load_workbook(config.COMBINED_EXCEL)
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            _style(ws, ws.max_column)
            for col in ws.columns:
                max_len = max(len(str(c.value or "")) for c in col)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 35)
        wb.save(config.COMBINED_EXCEL)

    # Eligible students
    eligible = get_all_eligible()
    if eligible:
        df_e = pd.DataFrame(eligible)
        df_e.columns = [c.replace("_", " ").title() for c in df_e.columns]
        with pd.ExcelWriter(config.ELIGIBLE_EXCEL, engine="openpyxl") as writer:
            df_e.to_excel(writer, sheet_name="Eligible Students", index=False)
        wb2 = load_workbook(config.ELIGIBLE_EXCEL)
        ws2 = wb2.active
        _style(ws2, ws2.max_column)
        for col in ws2.columns:
            max_len = max(len(str(c.value or "")) for c in col)
            ws2.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)
        wb2.save(config.ELIGIBLE_EXCEL)
